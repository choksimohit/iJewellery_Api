import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

rows = [
    ("Trusted Device Tokens", "GetUserByDeviceToken", "Looks up a user by device token in mchoksi.TrustedDevices and updates LastUsed timestamp"),
    ("Trusted Device Tokens", "RegisterTrustedDevice", "Upserts a trusted device record (token, userId, deviceName) — inserts if new, updates if exists"),
    ("Entity / Customer Photos", "UpdateEntityPhoto", "MERGE upserts a BLOB photo into mchoksi.m_EntityPhoto for any entity type with content-type tracking"),
    ("Entity / Customer Photos", "GetEntityPhoto", "Fetches photo bytes and content-type from mchoksi.m_EntityPhoto by entity type and ID"),
    ("Entity / Customer Photos", "UpdateCustomerPhoto", "Convenience wrapper — calls UpdateEntityPhoto with EntityType='Customer'"),
    ("Entity / Customer Photos", "GetCustomerPhoto", "Convenience wrapper — calls GetEntityPhoto with EntityType='Customer'"),
    ("Metal Rates", "UpdateMetalRates", "Inserts a new gold and silver rate record into mchoksi.MetalRates with timestamp and creator"),
    ("Metal Rates", "GetLatestMetalRates", "Returns the most recent gold and silver rate from mchoksi.MetalRates ordered by CreatedAt DESC"),
    ("Admin Menu Page CRUD", "GetAllMenuPages", "Fetches all rows from mchoksi.MenuPages ordered by CategoryOrder and SortOrder"),
    ("Admin Menu Page CRUD", "InsertMenuPage", "Inserts a new sidebar nav item with URL, display name, sort order, category, icons, and app version"),
    ("Admin Menu Page CRUD", "UpdateMenuPage", "Updates an existing menu page record by MenuID"),
    ("Admin Menu Page CRUD", "DeleteMenuPage", "Deletes a menu page record by MenuID"),
    ("User Authorization CRUD", "GetAllUserAuthorizations", "Fetches all rows from mchoksi.UserAuthorization ordered by UserID and PageName"),
    ("User Authorization CRUD", "InsertUserAuthorization", "Inserts a new user-page access record with CanAccess flag"),
    ("User Authorization CRUD", "UpdateUserAuthorization", "Updates an existing user authorization record by ID"),
    ("User Authorization CRUD", "DeleteUserAuthorization", "Deletes a user authorization record by ID"),
    ("Sidebar Menu Loader", "LoadMenuPages", "Calls iJewellery_Get_SidebarMenu SP for a given user and optional app version; fails silently"),
    ("Multi-item Loan Insert", "InsertLoan_Multi", "Inserts a loan header row within a shared transaction — used for multi-item loan creation"),
    ("Multi-item Loan Insert", "InsertLoanItem_Multi", "Inserts a single loan item row within a shared transaction with weight precision (3 decimal places)"),
    ("Multi-item Loan Insert", "DeleteLoanItems", "Deletes all item rows for a loan within a shared transaction via iJewellery_Delete_Loan_Items SP"),
    ("Loan Header Update", "UpdateLoanHeader", "Updates only the loan date and amount on a loan header via iJewellery_Update_Loan_Header SP"),
    ("Khatabook", "KhataAddBill", "Creates a new bill entry for a customer via usp_Khata_AddBill SP; returns the new BillID"),
    ("Khatabook", "KhataAddDebitToBill", "Adds a debit transaction to an existing bill via usp_Khata_AddDebitToBill SP"),
    ("Khatabook", "KhataAddPayment", "Records a payment against a bill via usp_Khata_AddPayment SP"),
    ("Khatabook", "KhataGetCustomerBalances", "Returns open/all customer balances via usp_Khata_GetCustomerBalances SP"),
    ("Khatabook", "KhataGetCustomerBills", "Returns all bills for a customer via usp_Khata_GetCustomerBills SP"),
    ("Khatabook", "KhataGetAllBillTransactions", "Returns all bill transactions for a customer via usp_Khata_GetAllBillTransactions SP"),
    ("Khatabook", "KhataGetOpenBillsForPayment", "Returns only open/unpaid bills for a customer via usp_Khata_GetOpenBillsForPayment SP"),
    ("Khatabook", "KhataDeleteTransaction", "Soft-deletes a transaction by TxnID via usp_Khata_DeleteTransaction SP"),
    ("Khatabook", "KhataDeleteBill", "Soft-deletes a bill by BillID via usp_Khata_DeleteBill SP"),
    ("Manage Customer (extended)", "GetCustomers", "Fetches all customers via usp_GetCustomers SP"),
    ("Manage Customer (extended)", "InsertCustomer", "Inserts a new customer and returns the new CustomerID output parameter via usp_InsertCustomer SP"),
    ("Manage Customer (extended)", "UpdateCustomer", "Updates name and address for an existing customer via usp_UpdateCustomer SP"),
    ("Manage Customer (extended)", "DeleteCustomer", "Deletes a customer record by CustomerID via usp_DeleteCustomer SP"),
    ("Manage Customer (extended)", "GetPhonesByCustomer", "Returns all phone numbers for a customer via usp_GetPhonesByCustomer SP"),
    ("Manage Customer (extended)", "InsertCustomerPhone", "Inserts a phone number for a customer and returns the new PhoneID via usp_InsertCustomerPhone SP"),
    ("Manage Customer (extended)", "DeleteCustomerPhone", "Deletes a customer phone record by CustomerPhoneID via usp_DeleteCustomerPhone SP"),
    ("Manage Customer (extended)", "SetPrimaryPhone", "Sets a specific phone as the primary contact for a customer via usp_SetPrimaryPhone SP"),
    ("Manage Customer (extended)", "MergeCustomers", "Merges duplicate customer IDs into a master customer via usp_MergeCustomers SP"),
    ("Borrowed Loans", "GetAllBorrowedLoans", "Returns all borrowed loans filtered by status (ALL/OPN/CLS) via iJewellery_Get_All_BorrowedLoans SP"),
    ("Borrowed Loans", "GetBorrowedLoanByID", "Returns a single borrowed loan record by BorrowedLoanID via iJewellery_Get_BorrowedLoan_ByID SP"),
    ("Borrowed Loans", "InsertBorrowedLoan", "Inserts a new borrowed loan with party details, interest rate/type, due date, and notes; returns new BorrowedLoanID"),
    ("Borrowed Loans", "UpdateBorrowedLoan", "Updates all fields of a borrowed loan record by BorrowedLoanID"),
    ("Borrowed Loans", "CloseBorrowedLoan", "Closes a borrowed loan with closure date, amount, and notes via iJewellery_Close_BorrowedLoan SP"),
    ("Borrowed Loans", "DeleteBorrowedLoan", "Soft-deletes a borrowed loan by BorrowedLoanID via iJewellery_Delete_BorrowedLoan SP"),
    ("Old Jewellery", "GetAllOldJewellery", "Fetches all old/second-hand jewellery purchase records via Ijewellery_GET_All_OldJewellery SP"),
    ("Old Jewellery", "InsertOldJewellery", "Inserts a second-hand jewellery purchase with metal type, rate, weight, price paid, and description"),
    ("Interest Dashboard", "GetInterestDashboardData", "Calls RPT_InterestEarnedFromClosure SP for interest earned dashboard reporting"),
    ("Barcode Test", "GetBarCodeTestData", "Calls Ijewellery_BarcodeTest SP — used for barcode printing/testing"),
    ("QR Code URL Update", "UpdateQRCodeURL", "Updates the QR code URL on a loan record via iJewellery_UpdateQRCodeURL SP"),
    ("Exception Logging", "ExceptionLog", "Logs application errors to DB via ACS_PHY_ExceptionLog SP, appending machine name to the message"),
    ("Decimal Rounding via SQL", "ConvertDecimalRound", "Rounds a decimal to 2 places using SQL CONVERT(DECIMAL(30,2)) for DB-consistent rounding"),
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Missing Functions"

thin = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap_top = Alignment(wrap_text=True, vertical="top")

# Header row
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", fgColor="2E4057")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

ws.append(["#", "Feature Group", "Method Name", "Description"])
for col in range(1, 5):
    cell = ws.cell(1, col)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = border
ws.row_dimensions[1].height = 22

# Group colour map
color_palette = ["EAF2FB", "E9F7EF", "FEF9E7", "FDEDEC", "F3E5F5", "E8F8F5", "FDF2E9", "EBF5FB", "F9EBEA", "E8DAEF"]
group_colors = {}
color_index = 0

for i, (group, method, desc) in enumerate(rows, 1):
    if group not in group_colors:
        group_colors[group] = color_palette[color_index % len(color_palette)]
        color_index += 1
    fill = PatternFill("solid", fgColor=group_colors[group])
    ws.append([i, group, method, desc])
    for col in range(1, 5):
        cell = ws.cell(i + 1, col)
        cell.fill = fill
        cell.alignment = wrap_top
        cell.border = border

# Column widths
ws.column_dimensions["A"].width = 5
ws.column_dimensions["B"].width = 28
ws.column_dimensions["C"].width = 34
ws.column_dimensions["D"].width = 72

wb.save("missing_functions.xlsx")
print("Saved: missing_functions.xlsx")
